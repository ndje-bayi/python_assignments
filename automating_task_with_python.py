from openpyxl import load_workbook
import csv

def main():

    try:

#Update the email addresses in excel format

        employees = Employees_Excel("employeedata.xlsx")

        print("Updating emails in .xlsx...")    

        employees.update_all_email_domaine("@handsinhands.org")

        print("Updated emails successfully \nSaving in excel file...")
        
        employees.save_data()
        
        print("Successfully saved\n\n")

#Update the email addresses in csv format

        employees = Employees_csv("employeedata.csv")

        print("Updating emails in .csv...")

        employees.update_all_email_domaine("@handsinhands.org")

        print("Updated emails successfully \nSaving in csv file...")

        employees.save_data()

        print("Successfully saved")

    except FileNotFoundError as e:
        print(e)


        
    
class Employees_Excel:

    def __init__(self, file_path:str, name_column:int=1, email_column:int=2,
                 phone_column:int=3):

        self.file_path = file_path

        #load data from excel file
        
        self.workbook = load_workbook(file_path)

        self.worksheet  = self.workbook.active

        #get the number of rows or entries in the excel file

        upper_bound = self.worksheet.max_row + 1

        #arrange the data in a suitable phython data structure
        
        self.data = []

        for i in range(2, upper_bound):
            
            name = self.worksheet.cell(row=i, column=name_column).value
            
            email = self.worksheet.cell(row=i, column=email_column).value
            
            phone = self.worksheet.cell(row=i, column=phone_column).value

            #ignore empty rows

            if not name and not email and not phone:
                continue

            self.data.append({"name":name, "email":email, "phone":phone})



    def update_all_email_domaine(self, new_domaine:str):
        """Updates the domaine name for all employees in the object's data"""
    
        for employee in self.data:
            
            current_email = employee["email"]

            name = current_email.split("@")[0] 

            employee["email"] = name + new_domaine

            
        
    def save_data(self):
        """saves back the object's data in the original excel file format"""

        for i, employee in enumerate(self.data, 2):
            
            for c, value in enumerate(employee.values(), 1):

                self.worksheet.cell(row=i, column=c).value = value

        self.workbook.save(self.file_path)
                
class Employees_csv:

    def __init__(self, file_path:str, name_index:int=0, email_index:int=1,
                 phone_index:int=2):
        
        self.file_path = file_path

        self.name_index = name_index

        self.email_index = email_index

        self.phone_index = phone_index

        #load data from csv file and
        #arrange it in a suitable python data structure

        with open(file_path, newline="") as file:
            
            reader = csv.reader(file)
            
            self.header = next(reader)

            self.data = [row for row in reader]


    def update_all_email_domaine(self, new_domaine:str):
        """Updates the domaine name for all employees in the object's data"""

        for employee in self.data:

            name = employee[self.email_index].split("@")[0]

            employee[self.email_index] = name + new_domaine


    def save_data(self):
        """saves back the object's data in the original excel file format"""

        with open(self.file_path, "w") as file:
            
            writer = csv.writer(file, lineterminator="\n")
            
            writer.writerow(self.header)

            for row in self.data:
                
                writer.writerow(row)
    

if __name__ == "__main__":
    main()
